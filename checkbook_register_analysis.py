"""
FILE: checkbook_register_analysis.py

PURPOSE: 

AUTHOR: 
Clinton G. 

TODO: Fill out this doc string
TODO: This algorithm needs to be adjusted
		so that only the current year is considered
		when parsing the personal register.
TODO:  Need to look for duplicates in transactions
"""
import sys
import re
import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import logging

# ----------------------------------------------------------------------- #
# Regular Expression Strings
# ----------------------------------------------------------------------- #


## HARDCODED COLUMN POSITIONS ##
# -------------------------------------- #
# Hardcoded column Positions
# -------------------------------------- #
col_position_personal_month = 3
col_position_personal_value	= 8
col_position_personal_des	= 6

col_position_bank_date 		= 1
col_position_bank_value		= 5
col_position_bank_des 		= 3

# -------------------------------------- #
# Dictionaries
# -------------------------------------- #
dict_bank_register	= {}
dict_personal_register	= {}


# -------------------------------------- #
# Boolean Flags
# -------------------------------------- #

# -------------------------------------- #
# Lists
# -------------------------------------- #
values 						= []
descriptions   				= []
months   					= []

# -------------------------------------- #
# Local Methods
# -------------------------------------- #
def debug_while():
	while(1):
		pass

def pause():
	user_input=input("Press any key to exit...")
	sys.exit(0)


# ------------------------------------- #
# Setup Logging
# -------------------------------------- #
logging.basicConfig(
	filename = 'register_balance_analysis.log',
	level = logging.DEBUG,
	format =' %(asctime)s -  %(levelname)s - %(message)s',
	filemode = 'w'
)


#****************************************************************************** 
#******************************  ---MAIN---  **********************************
#******************************************************************************   
if __name__ == '__main__':

	# ----------------------------------------------------------------------- #
	# Iterate through files and delete
	# existing comparison BOMs and log files if they exist
	# ----------------------------------------------------------------------- #

	path = os.getcwd()
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files
	
	for i in range(len(files)):
		if(files[i].upper().find("ANALYSIS") != -1 and files[i].upper().find("REGISTER") != -1
	  		and not files[i].upper().endswith(".PY") and not files[i].upper().endswith(".LOG")
			and not files[i].upper().endswith(".MD")) :
			os.remove(files[i])
		
		if(files[i].upper().endswith(".CSV")):
			# ----------------------------------------------------------------------- #
			# The bank file will download as a csv, so it needs to be converted to 
			# an Excel file
			# ----------------------------------------------------------------------- #
			csvwb = Workbook()
			csvws = csvwb.active
			csvws.title = "Bank Register"
			
			with open(files[i]) as f:
				reader = csv.reader(f, delimiter=',')
				for row in reader:
					csvws.append(row)

				csvwb.save(filename="bank_register.xlsx")
	 
	# ----------------------------------------------------------------------- #
	# Some file may have been removed, so refresh 
	# directory information.  
	# ----------------------------------------------------------------------- #
	path = os.getcwd()
	for (path, dirs, files) in os.walk(path):
		path
		dirs
		files

	print ("Files found in directory: ", str(len(files)))
	logging.info("Files found in directory: " + str(len(files)))
	print ("File names: ", files)
	for i in range(len(files)):
		logging.info("File " + str(i+1) + ") " + files[i])


	# ----------------------------------------------------------------------- #
	# Iterate through files
	# ----------------------------------------------------------------------- #
	for i in range(len(files)):
		
		# ----------------------------------------------------------------------- #
		# Only open files having the proper extension 
		# ----------------------------------------------------------------------- #
		if(files[i].upper().endswith(".XLSX")):
			
			print ("\n===============================================")
			print ("===============================================")
			print ("Opening file: ", files[i])

			logging.info ("===============================================")
			logging.info ("===============================================")
			logging.info ("Opening file: " + files[i])
			
			# ----------------------------------------------------------------------- #
			# Define the register type
			# ----------------------------------------------------------------------- #
			user_input = str(input("Is this a personal workbook (opposed to the export from the bank) (y/n): "))
			
			if(user_input == 'y' or user_input == 'Y'):
				register_type_is_personal = True
				logging.info ("This register is the personal register")
			else:
				register_type_is_personal = False
				logging.info ("This register was exported from the bank")
				

			wb = load_workbook(filename = files[i])     # Open the workbook that we are going to parse though 
			ws = wb.sheetnames             				# Grab the names of the worksheets -- I believe this line is critical.
			
			# There shall be only one sheet in each workbook
			num_sheets = len(ws)						# This is the number of sheets
			if(num_sheets > 1):
				print("The workbook contains more than on sheet!")
				logging.info("   ***The workbook contains more than on sheet!")
			

			print ("The number of worksheets in this book: ", str(num_sheets))
			print ("Worksheet names: ", ws)
			print ("===============================================")
			
			logging.info ("The number of worksheets in this book: " + str(num_sheets))
			for i in range (len(ws)):
				logging.info ("Worksheet " + str(i) + ") " + str(ws))

			
			# ----------------------------------------------------------------------- #
			# Ask the user what month we want to start with 
			# ----------------------------------------------------------------------- #
			start_month = 99
			while(start_month < 1 or start_month > 12):
				print("What month do we want to start with (i.e. JAN = 1)");
				start_month=int(input("Enter the start month: "))

				if(start_month < 1 or start_month > 12):
					print("   ***Incorrect start month entered for start month")
			
			
			print("The start month entered: ", str(start_month))
			logging.info("The start month entered: " + str(start_month))
			
			current_sheet = wb[ws[0]]

			# ----------------------------------------------------------------------- #
			# Iterate through every row on current sheet
			# ----------------------------------------------------------------------- #
			num_rows = current_sheet.max_row     		
			num_cols = current_sheet.max_column 
			for current_row in range (1,num_rows + 1):					# The methods start at 1, not 0
			
				if(register_type_is_personal):
					# ----------------------------------------------------------------------- #
					# Grab the month number
					# ----------------------------------------------------------------------- #
					month_number = int(current_sheet.cell(row = current_row, column=col_position_personal_month).value) 
					logging.info("Month number from personal register: " + str(month_number)) 

					# ----------------------------------------------------------------------- #
					# Grab the transaction number
					# ----------------------------------------------------------------------- #
					transaction_amount = float(current_sheet.cell(row = current_row, column=col_position_personal_value).value)

					# ----------------------------------------------------------------------- #
					# Grab the description
					# ----------------------------------------------------------------------- #
					transaction_description = str(current_sheet.cell(row = current_row, column=col_position_personal_des).value)


				else:
					
					# ----------------------------------------------------------------------- #
					# Grab the month number
					# ----------------------------------------------------------------------- #
					# month_number = str(current_sheet.cell(row = current_row, column=col_position_bank_date ).value).encode(encoding = 'UTF-8') 
					month_number = str(current_sheet.cell(row = current_row, column=col_position_bank_date ).value) 
					month_number = month_number.split('/')
					month_number = int(month_number[0].strip())
					logging.info("Month number from bank generated register: " + str(month_number)) 

					# ----------------------------------------------------------------------- #
					# Grab the transaction number
					# ----------------------------------------------------------------------- #
					transaction_amount = float(current_sheet.cell(row = current_row, column=col_position_bank_value).value)
					# transaction_amount = current_sheet.cell(row = current_row, column=col_position_bank_value).value
					
					# ----------------------------------------------------------------------- #
					# Grab the description
					# ----------------------------------------------------------------------- #
					transaction_description = str(current_sheet.cell(row = current_row, column=col_position_bank_des).value)

				# ----------------------------------------------------------------------- #
				# If the month number is correct, append values to lists
				# ----------------------------------------------------------------------- #
				if(month_number >= start_month):
					values.append(transaction_amount)
					months.append(month_number)
					descriptions.append(transaction_description)

			#########
			# END iterating over rows
			#########

			# ---------------------------------------------------------------------- #
			# Build the dictionary  
			# ----------------------------------------------------------------------- #
			if(register_type_is_personal):
				for i in range (0,len(values)):				
					dict_personal_register[values[i]] = (descriptions[i],months[i])
			else:
				for i in range (0,len(values)):				
					dict_bank_register[values[i]] = (descriptions[i],months[i])

			#----------------------------------------------------------------------- #
			# Lists shall be cleared before moving onto the 
			# next file, as a different dictionary will need to populated
			#----------------------------------------------------------------------- #
			values.clear()
			descriptions.clear()
			months.clear()

		#########
		# END iterating over files
		#########
			
	# ----------------------------------------------------------------------- #
	# Iterate through every transaction in the personal register 
	# so we can compare against the bank register
	# It also needs to be verified that the number of transactions are the same
	# for example, if there are four $100 transactions in the personal register
	# there there shall be four $100 transactions in the bank's register
	# ----------------------------------------------------------------------- #
	print("\n\n>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
	print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
	
	print("Transactions in PERSONAL but NOT BANK")
	for key in dict_personal_register:
		if(key in dict_bank_register):
			print(".", end=' ')
		if (key not in dict_bank_register):
			print("\nAmount:\t\t ", key)
			print("Description:\t ", str(dict_personal_register[key][0]))

	print("\n\nTransactions in BANK but NOT PERSONAL")
	for key in dict_bank_register:
		if(key in dict_personal_register):
			print(".",end=' ')
		if (key not in dict_personal_register):
			print("\nAmount: ", key)
			print("Description:\t ", str(dict_bank_register[key][0]))

	logging.info("Processing complete")


	null=input("\n\nPress any key to close...")